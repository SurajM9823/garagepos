from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from garage.models import VehicleCompany, VehicleModel, VehicleType, ClientGarage
import openpyxl
from openpyxl.utils import get_column_letter
from django.views.decorators.csrf import csrf_exempt
import json

@login_required
def admin_upload(request):
    if request.user.is_superuser or request.user.role != 'admin':
        return redirect(f'/{request.user.role}/dashboard/')
    return render(request, 'admin/upload.html', {'user': request.user})

@login_required
def get_companies(request):
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 5))
    companies = VehicleCompany.objects.filter(client_garage=request.user.client_garage)
    paginator = Paginator(companies, per_page)
    page_obj = paginator.get_page(page)
    data = [{'id': c.id, 'name': c.name, 'description': c.description} for c in page_obj]
    return JsonResponse({'items': data, 'total': paginator.count})

@login_required
def get_vehicle_types(request):
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 5))
    vehicle_types = VehicleType.objects.filter(client_garage=request.user.client_garage)
    paginator = Paginator(vehicle_types, per_page)
    page_obj = paginator.get_page(page)
    data = [{'id': t.id, 'name': t.name, 'description': t.description} for t in page_obj]
    return JsonResponse({'items': data, 'total': paginator.count})

@login_required
def get_models(request):
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 5))
    models = VehicleModel.objects.filter(client_garage=request.user.client_garage)
    paginator = Paginator(models, per_page)
    page_obj = paginator.get_page(page)
    data = [{'id': m.id, 'name': m.name, 'company': m.company.id, 'company_name': m.company.name, 'vehicle_type': m.vehicle_type.id, 'vehicle_type_name': m.vehicle_type.name, 'description': m.description} for m in page_obj]
    return JsonResponse({'items': data, 'total': paginator.count})

@login_required
def get_company(request, id):
    company = VehicleCompany.objects.get(id=id, client_garage=request.user.client_garage)
    return JsonResponse({'id': company.id, 'name': company.name, 'description': company.description})

@login_required
def get_vehicle_type(request, id):
    vehicle_type = VehicleType.objects.get(id=id, client_garage=request.user.client_garage)
    return JsonResponse({'id': vehicle_type.id, 'name': vehicle_type.name, 'description': vehicle_type.description})

@login_required
def get_model(request, id):
    model = VehicleModel.objects.get(id=id, client_garage=request.user.client_garage)
    return JsonResponse({'id': model.id, 'name': model.name, 'company': model.company.id, 'vehicle_type': model.vehicle_type.id, 'description': model.description})

@csrf_exempt
@login_required
def save_companies(request, id=None):
    data = json.loads(request.body)
    if id:
        company = VehicleCompany.objects.get(id=id, client_garage=request.user.client_garage)
        company.name = data['name']
        company.description = data.get('description')
        company.save()
    else:
        VehicleCompany.objects.create(
            client_garage=request.user.client_garage,
            name=data['name'],
            description=data.get('description')
        )
    return JsonResponse({'status': 'success'})

@csrf_exempt
@login_required
def save_vehicle_types(request, id=None):
    data = json.loads(request.body)
    if id:
        vehicle_type = VehicleType.objects.get(id=id, client_garage=request.user.client_garage)
        vehicle_type.name = data['name']
        vehicle_type.description = data.get('description')
        vehicle_type.save()
    else:
        VehicleType.objects.create(
            client_garage=request.user.client_garage,
            name=data['name'],
            description=data.get('description')
        )
    return JsonResponse({'status': 'success'})

@csrf_exempt
@login_required
def save_models(request, id=None):
    data = json.loads(request.body)
    if id:
        model = VehicleModel.objects.get(id=id, client_garage=request.user.client_garage)
        model.name = data['name']
        model.company = VehicleCompany.objects.get(id=data['company'], client_garage=request.user.client_garage)
        model.vehicle_type = VehicleType.objects.get(id=data['vehicle_type'], client_garage=request.user.client_garage)
        model.description = data.get('description')
        model.save()
    else:
        VehicleModel.objects.create(
            client_garage=request.user.client_garage,
            name=data['name'],
            company=VehicleCompany.objects.get(id=data['company'], client_garage=request.user.client_garage),
            vehicle_type=VehicleType.objects.get(id=data['vehicle_type'], client_garage=request.user.client_garage),
            description=data.get('description')
        )
    return JsonResponse({'status': 'success'})

@csrf_exempt
@login_required
def delete_companies(request, id):
    company = VehicleCompany.objects.get(id=id, client_garage=request.user.client_garage)
    company.delete()
    return JsonResponse({'status': 'success'})

@csrf_exempt
@login_required
def delete_vehicle_types(request, id):
    vehicle_type = VehicleType.objects.get(id=id, client_garage=request.user.client_garage)
    vehicle_type.delete()
    return JsonResponse({'status': 'success'})

@csrf_exempt
@login_required
def delete_models(request, id):
    model = VehicleModel.objects.get(id=id, client_garage=request.user.client_garage)
    model.delete()
    return JsonResponse({'status': 'success'})

@login_required
def export_companies(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="companies.csv"'
    writer = openpyxl.Workbook()
    ws = writer.active
    ws.append(['Name', 'Description'])
    for company in VehicleCompany.objects.filter(client_garage=request.user.client_garage):
        ws.append([company.name, company.description or ''])
    writer.save(response)
    return response

@login_required
def export_vehicle_types(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="vehicle_types.csv"'
    writer = openpyxl.Workbook()
    ws = writer.active
    ws.append(['Name', 'Description'])
    for vehicle_type in VehicleType.objects.filter(client_garage=request.user.client_garage):
        ws.append([vehicle_type.name, vehicle_type.description or ''])
    writer.save(response)
    return response

@login_required
def export_models(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="models.csv"'
    writer = openpyxl.Workbook()
    ws = writer.active
    ws.append(['Name', 'Company', 'Vehicle Type', 'Description'])
    for model in VehicleModel.objects.filter(client_garage=request.user.client_garage):
        ws.append([model.name, model.company.name, model.vehicle_type.name, model.description or ''])
    writer.save(response)
    return response

@csrf_exempt
@login_required
def upload_companies(request):
    file = request.FILES['file']
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        VehicleCompany.objects.create(
            client_garage=request.user.client_garage,
            name=row[0],
            description=row[1] if len(row) > 1 else ''
        )
    return JsonResponse({'status': 'success'})

@csrf_exempt
@login_required
def upload_vehicle_types(request):
    file = request.FILES['file']
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        VehicleType.objects.create(
            client_garage=request.user.client_garage,
            name=row[0],
            description=row[1] if len(row) > 1 else ''
        )
    return JsonResponse({'status': 'success'})

@csrf_exempt
@login_required
def upload_models(request):
    file = request.FILES['file']
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            company = VehicleCompany.objects.get(name=row[1], client_garage=request.user.client_garage)
            vehicle_type = VehicleType.objects.get(name=row[2], client_garage=request.user.client_garage)
            VehicleModel.objects.create(
                client_garage=request.user.client_garage,
                name=row[0],
                company=company,
                vehicle_type=vehicle_type,
                description=row[3] if len(row) > 3 else ''
            )
        except (VehicleCompany.DoesNotExist, VehicleType.DoesNotExist):
            continue
    return JsonResponse({'status': 'success'})

@login_required
def download_template(request, type):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{type}_template.csv"'
    writer = openpyxl.Workbook()
    ws = writer.active
    if type == 'companies':
        ws.append(['Name', 'Description'])
        ws.append(['Example Company', 'Sample description'])
    elif type == 'vehicle-types':
        ws.append(['Name', 'Description'])
        ws.append(['Example Type', 'Sample description'])
    elif type == 'models':
        ws.append(['Name', 'Company', 'Vehicle Type', 'Description'])
        ws.append(['Example Model', 'Example Company', 'Example Type', 'Sample description'])
    writer.save(response)
    return response